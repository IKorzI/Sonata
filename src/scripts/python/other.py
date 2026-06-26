from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook
import os
from datetime import datetime
import time
import win32clipboard as cb
from PIL import Image
from datetime import datetime, timedelta

from utils import MONTH_NAMES_CAPITAL
from utils import save_file


def other_extend_image(path, final_path):
    """
    Extends the image by adding white margins at the edges.

    Args:
        path (str): Path to the source image.
        final_path (str): Path to save the processed image.
    Returns:
        bool: True after successful saving.
    """
    margin_percent = 2
    color = (255, 255, 255)

    img = Image.open(path)
    w, h = img.size

    # Calculation of margin sizes proportionally to the image dimensions
    border_x = int(w * margin_percent / 100)
    border_y = int(h * margin_percent / 100)
    border = min(border_x, border_y)

    # Creating a new image with a white background and pasting the original in the center
    new_img = Image.new("RGB", (w + border * 2, h + border * 2), color)
    new_img.paste(img, (border, border))

    new_img.save(final_path, "PNG")
    return True


def get_calendar_schedule(start_str, end_str, align_weekday=None):
    """
    Creates a calendar schedule of working days between two dates.

    Args:
        start_str (str): Start date in 'DD.MM.YYYY' format.
        end_str (str): End date in 'DD.MM.YYYY' format.
        align_weekday (int, optional): Day of the week for alignment (0=Mon...4=Fri). Defaults to None.
    Returns:
        dict: Dictionary with the calendar and the number of the first day.
    """

    start_date = datetime.strptime(start_str, "%d.%m.%Y")
    end_date = datetime.strptime(end_str, "%d.%m.%Y")

    all_work_days = []
    current_date = start_date

    # Collecting all working days (Monday to Friday)
    while current_date <= end_date:
        if current_date.weekday() < 5:
            all_work_days.append(current_date)
        current_date += timedelta(days=1)

    # Aligning the first week using empty values (None) for the shift
    if align_weekday is not None and all_work_days:
        first_real_weekday = all_work_days[0].weekday()

        offset = (first_real_weekday - align_weekday) % 5

        for _ in range(offset):
            all_work_days.insert(0, None)

    days = {}

    # Breaking down collected days into weeks (5 days each) and grouping them by months
    for i in range(0, len(all_work_days), 5):
        week_chunk = all_work_days[i : i + 5]

        first_real_day = next((d for d in week_chunk if d is not None), None)

        if first_real_day:
            month_key = f"{first_real_day.month:02d}"

            week_data = [(d.day if d else None) for d in week_chunk]

            while len(week_data) < 5:
                week_data.append(None)

            if month_key not in days:
                days[month_key] = []
            days[month_key].append(week_data)

    # Determining the number of the first day of the week for a correct starting count
    if all_work_days:
        if align_weekday is not None:
            first_day_num = align_weekday + 1
        else:
            first_real_obj = next((d for d in all_work_days if d is not None), None)
            if first_real_obj:
                first_day_num = first_real_obj.weekday() + 1
            else:
                first_day_num = 1
    else:
        first_day_num = start_date.weekday() + 1

    return {"calendar": days, "first_day": first_day_num}


def other_NumDenStart(info, app_path):
    """
    Creates and fills the numerator and denominator file for two semesters.

    Args:
        info (dict): Information about semester dates and the file path.
        app_path (str): Path to the application.
    Returns:
        dict: Result of the operation with the list of created files.
    """
    answer = {"success": True, "files": []}
    path_to_save = os.path.dirname(info["file_path"])
    os.makedirs(path_to_save, exist_ok=True)
    path = f"{app_path}/public/examples/work/"
    days_of_week = ["ПН", "ВВ", "СР", "ЧТ", "ПТ"]

    workbook = load_workbook(f"{path}/num-den.xlsx")
    gray_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # =============================================================================================================
    # FIRST SEMESTER

    sheet = workbook["I семестр"]

    result1 = get_calendar_schedule(info["semester_1_start"], info["semester_1_end"])

    sem_1_weekday_idx = result1["first_day"] - 1

    # Filling the table header with days of the week, taking into account the calculated start day
    col_num = 5
    for i in range(result1["first_day"], 5 + 1):
        sheet.cell(row=4, column=col_num).value = days_of_week[i - 1]
        col_num += 1
    for i in range(1, result1["first_day"]):
        sheet.cell(row=4, column=col_num).value = days_of_week[i - 1]
        col_num += 1

    week_num = 1
    step = 0
    dels = 0

    # Iterating through months and their weeks to fill the table
    for month_num, calendar in result1["calendar"].items():
        row = 5 + 7 * step - dels
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        sheet.cell(row=row, column=3).value = MONTH_NAMES_CAPITAL[month_num]
        for i, weak in enumerate(calendar):
            row = (5 + 7 * step) + i + 1 - dels
            sheet.cell(row=row, column=3).value = week_num

            # Determining the week's parity (Numerator/Denominator) and applying styles
            if week_num % 2 != 0:
                sheet.cell(row=row, column=4).value = "Чисельник"
                sheet.cell(row=row, column=4).alignment = Alignment(
                    horizontal="left", vertical="center"
                )
            else:
                sheet.cell(row=row, column=4).value = "Знаменник"
                sheet.cell(row=row, column=4).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                # Coloring the row gray for the "Denominator"
                for col_num in range(3, 9 + 1):
                    sheet.cell(row=row, column=col_num).fill = gray_fill

            week_num += 1
            for a, date in enumerate(weak):
                if date is not None:
                    sheet.cell(row=row, column=4 + a + 1).value = date

        # Deleting extra rows if a month has fewer than 6 weeks
        if len(calendar) < 6:
            for i in range(len(calendar) + 1, 6 + 1):
                row = (5 + 7 * step) + i - dels
                sheet.delete_rows(row)
                dels += 1
        step += 1

    # Hiding empty rows at the end of the sheet
    for row in range(35 - dels, 35):
        sheet.row_dimensions[row].hidden = True

    # =============================================================================================================
    # SECOND SEMESTER

    sheet = workbook["II семестр"]

    result2 = get_calendar_schedule(
        info["semester_2_start"],
        info["semester_2_end"],
        align_weekday=sem_1_weekday_idx,
    )

    col_num = 5
    for i in range(result1["first_day"], 5 + 1):
        sheet.cell(row=4, column=col_num).value = days_of_week[i - 1]
        col_num += 1
    for i in range(1, result1["first_day"]):
        sheet.cell(row=4, column=col_num).value = days_of_week[i - 1]
        col_num += 1

    s1_date = datetime.strptime(info["semester_1_start"], "%d.%m.%Y")
    s2_date = datetime.strptime(info["semester_2_start"], "%d.%m.%Y")

    monday_sem1 = s1_date - timedelta(days=s1_date.weekday())
    monday_sem2 = s2_date - timedelta(days=s2_date.weekday())

    # Calculating the number of weeks that have passed between semesters to maintain the correct parity
    delta_days = (monday_sem2 - monday_sem1).days
    weeks_passed = delta_days // 7
    current_parity_week = 1 + weeks_passed

    display_week_num = 1
    step = 0
    dels = 0

    # Filling dates for the second semester similarly to the first, but taking into account the global week parity
    for month_num, calendar in result2["calendar"].items():
        row = 5 + 7 * step - dels
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        sheet.cell(row=row, column=3).value = MONTH_NAMES_CAPITAL[month_num]

        for i, weak in enumerate(calendar):
            row = (5 + 7 * step) + i + 1 - dels
            sheet.cell(row=row, column=3).value = display_week_num

            if current_parity_week % 2 != 0:
                sheet.cell(row=row, column=4).value = "Чисельник"
                sheet.cell(row=row, column=4).alignment = Alignment(
                    horizontal="left", vertical="center"
                )
            else:
                sheet.cell(row=row, column=4).value = "Знаменник"
                sheet.cell(row=row, column=4).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                for col_num in range(3, 9 + 1):
                    sheet.cell(row=row, column=col_num).fill = gray_fill

            current_parity_week += 1
            display_week_num += 1

            for a, date in enumerate(weak):
                if date is not None:
                    sheet.cell(row=row, column=4 + a + 1).value = date

        if len(calendar) < 6:
            for i in range(len(calendar) + 1, 6 + 1):
                row = (5 + 7 * step) + i - dels
                sheet.delete_rows(row)
                dels += 1
        step += 1

    for row in range(49 - dels, 49):
        sheet.row_dimensions[row].hidden = True

    workbook_path = save_file(workbook, info["file_path"])
    if workbook_path != True:
        answer["files"].append(workbook_path)

    return answer
