from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

from utils import MONTH_NAMES, FILL_GRAY
from utils import cleaning_and_save_workbook

def filling_out_the_hours_sheet(sheet: Worksheet, month_name: str, year: int, group: object, days: object):
    """
    Fills out the hours sheet.

    Args:
        sheet (Worksheet): The sheet to be filled out.
        month_name (str): The name of the month in lowercase.
        year (int): The year.
        group (object): The group of students.
        days (object): The list of days.
    Returns:
        sheet (Worksheet): The filled out sheet.
    """
    subject_len = len(group['subjects'])
    days_list = list(range(days['start_day'], days['end_day'] + 1))
    start_weekday = days['start_weekday']

    # Calculating bounds to delete extra rows
    delete_row_start = 11 + subject_len + 1
    delete_row_count = 42 - delete_row_start
    end_row = delete_row_start - 1
    
    # Calculating bounds to delete extra columns
    delete_col_count = (days['start_day'] - 1) + (31 - days['end_day'])
    delete_col_start = 36 - delete_col_count + 1
    end_col = delete_col_start - 1
    end_col_letter = get_column_letter(end_col)
    end_col_plus_1_letter = get_column_letter(end_col + 1)

    dims = sheet.column_dimensions

    # =============================================================================================================
    # PREPARING THE SHEET BEFORE FILLING

    sheet.unmerge_cells(start_row=45, start_column=3, end_row=45, end_column=38)
    sheet.delete_rows(delete_row_start, amount=delete_row_count)
    for row in range(delete_row_start + 6, delete_row_start + 6 + delete_row_count):
        sheet.row_dimensions[row].hidden = True
    sheet.merge_cells(start_row=delete_row_start + 3, start_column=3, end_row=delete_row_start + 3, end_column=38)

    # Adjusting the sheet structure if there are extra days (columns)
    if delete_col_count > 0:
        sheet.unmerge_cells(start_row=10, start_column=6, end_row=10, end_column=36)
        for row in [5, 6, 7, 8, 9, end_row + 4]:
            sheet.unmerge_cells(start_row=row, start_column=3, end_row=row, end_column=38)
        for col in [37, 38]:
            sheet.unmerge_cells(start_row=10, start_column=col, end_row=11, end_column=col)
        sheet.delete_cols(idx=delete_col_start, amount=delete_col_count)
        for col in range(delete_col_start + 4, delete_col_start + 4 + delete_col_count):
            dims[get_column_letter(col)].hidden = True
        sheet.merge_cells(start_row=10, start_column=6, end_row=10, end_column=delete_col_start - 1)
        for row in [5, 6, 7, 8, 9, end_row + 4]:
            sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=delete_col_start + 1)
        for col in [delete_col_start, delete_col_start + 1]:
            sheet.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)
        dims[get_column_letter(delete_col_start)].width = 4.05
        dims[get_column_letter(delete_col_start + 1)].width = 5.55
        dims[get_column_letter(delete_col_start + 2)].width = 1.5
        dims[get_column_letter(delete_col_start + 3)].width = 1.5
    
    col_letter = get_column_letter(end_col + 1)
    sheet.print_area = f'C3:{col_letter}{end_row + 4}'
    
    # Adding formulas to calculate worked hours by columns and rows
    for col in range(6, end_col + 1):
        col_letter = get_column_letter(col)
        sheet.cell(row=end_row + 1, column=col).value = f'=IF(${end_col_plus_1_letter}{end_row + 1}="Очікую","Оч",SUM({col_letter}12:{col_letter}{end_row}))'
    for row in range(12, end_row + 1):
        sheet.cell(row=row, column=end_col + 1).value = f'=IF(SUM(F{row}:{end_col_letter}{row})=0,"Очікую",SUM(F{row}:{end_col_letter}{row}))'
    sheet.cell(row=end_row + 1, column=end_col + 1).value = f'=IF(SUM({end_col_plus_1_letter}12:{end_col_plus_1_letter}{end_row})=0,"Очікую",SUM({end_col_plus_1_letter}12:{end_col_plus_1_letter}{end_row}))'

    # =============================================================================================================
    # FILLING THE SHEET

    sheet.cell(row=7, column=3).value = f'за {month_name} {year} р.'
    sheet.cell(row=8, column=3).value = group['cell_C8']

    # Filling dates in the header
    for index, date in enumerate(days_list):
        sheet.cell(row=11, column=index + 6).value = date
    
    # Filling subjects, teachers, and worked hours
    for index, subject in enumerate(group['subjects']):
        row = index + 12
        sheet.cell(row=row, column=4).value = f'{subject['subject_name']}'
        sheet.cell(row=row, column=5).value = f'{subject['teacher_name']}'
        for index, hour in enumerate(subject['values']):
            if hour:
                sheet.cell(row=row, column=index + 6).value = hour
        
        # Identifying weekends and coloring them
        weekends = [val for i, val in enumerate(days_list) if (i + start_weekday + 1) % 7 in [6, 0]]
        for date in weekends:
            sheet.cell(row=row, column=5 + (date - days_list[0] + 1)).fill = FILL_GRAY

def filling_out_the_timekeeping_sheet(workbook, months, hours_per_subject):
    """
    Fills out the timekeeping sheet.

    Args:
        workbook (Workbook): The workbook.
        months (list): The list of months.
        hours_per_subject (dict): Hours per subject.
    Returns:
        dict: A dictionary with the updated workbook and found errors.
    """
    delete_col_start = 6 + len(months) + 1
    delete_col_count = 19 - delete_col_start
    end_col = delete_col_start - 1
    end_col_letter = get_column_letter(end_col)
    end_col_plus_1_letter = get_column_letter(end_col + 1)

    not_found_subjects = []

    step = 1
    for group_index, group in enumerate(months[0]['groups']):

        group_code = group['group_code']

        delete_row_start = 3 + len(group['subjects']) + 1
        delete_row_count = 34 - delete_row_start
        end_row = delete_row_start - 1

        sheet = workbook[f'Л{step}']
        sheet.title = group_code
        dims = sheet.column_dimensions

    # =============================================================================================================
    # PREPARING THE SHEET

        sheet.unmerge_cells(start_row=34, start_column=3, end_row=34, end_column=5)
        sheet.delete_rows(delete_row_start, amount=delete_row_count)
        for row in range(delete_row_start + 3, delete_row_start + 3 + delete_row_count):
            sheet.row_dimensions[row].hidden = True
        sheet.merge_cells(start_row=delete_row_start, start_column=3, end_row=delete_row_start, end_column=5)

        # Deleting extra columns if the number of months is less than the maximum limit
        if delete_col_count > 0:
            sheet.delete_cols(idx=delete_col_start, amount=delete_col_count)
            for col in range(delete_col_start + 3, delete_col_start + 3 + delete_col_count):
                dims[get_column_letter(col)].hidden = True
        dims[get_column_letter(end_col + 2)].width = 2.71
        dims[get_column_letter(end_col + 3)].width = 2.71

        sheet.print_area = f'C3:{end_col_plus_1_letter}{end_row + 1}'

        # Adding formulas to check discrepancies (surplus/shortage of hours)
        for row in range(4, end_row + 1):
            sheet.cell(row=row, column=end_col + 1).value = f'=IF(COUNTBLANK(F{row}:{end_col_letter}{row})>0,"Очікую", IF(F{row}-SUM(G{row}:{end_col_letter}{row})=0,"0", IF(F{row}-SUM(G{row}:{end_col_letter}{row})<0,"Надлишок " & (F{row}-SUM(G{row}:{end_col_letter}{row}))*(-1),"Не вистачає " & F{row}-SUM(G{row}:{end_col_letter}{row}))))'
        for col in range(6, end_col + 1):
            col_letter = get_column_letter(col)
            sheet.cell(row=end_row + 1, column=col).value = f'=IF(COUNTBLANK({col_letter}4:{col_letter}{end_row})>0,"Очікую",SUM({col_letter}4:{col_letter}{end_row}))'
        sheet.cell(row=end_row + 1, column=end_col + 1).value = f'=IF(COUNTIF(F{end_row + 1}:{end_col_letter}{end_row + 1},"Очікую")>0,"Очікую", IF(F{end_row + 1}-SUM(G{end_row + 1}:{end_col_letter}{end_row + 1})=0,"0", IF(F{end_row + 1}-SUM(G{end_row + 1}:{end_col_letter}{end_row + 1})<0,"Надлишок "&(F{end_row + 1}-SUM(G{end_row + 1}:{end_col_letter}{end_row + 1}))*(-1),"Не вистачає "&F{end_row + 1}-SUM(G{end_row + 1}:{end_col_letter}{end_row + 1}))))'
        
    # =============================================================================================================
    # FILLING THE SHEET

        # Setting month names in the header
        col = 7
        for month in months:
            sheet.cell(row=3, column=col).value = month['month_name'].capitalize()
            col += 1
            
        # Filling subjects and teachers
        for index, subject in enumerate(group['subjects']):
            row = 4 + index

            sheet.cell(row=row, column=4).value = subject['subject_name']
            sheet.cell(row=row, column=5).value = subject['teacher_name']
            sheet.cell(row=row, column=6).value = hours_per_subject[subject['subject_name']]

            col = 7
            # Creating external links to specific month files
            for month in months:
                month_name = month['month_name']

                current_group = month['groups'][group_index]
                first_subject = current_group['subjects'][0]
                col_num = 5 + len(first_subject['values']) + 1
                col_letter = get_column_letter(col_num)

                sheet.cell(row=row, column=col).value = f"='[{month_name.capitalize()}.xlsx]{group_code}'!${col_letter}{8 + row}"
                col += 1

        step += 1

    return {
        'workbook': workbook,
        'not_found_subjects': not_found_subjects
    }

def filling_out_the_general_statement_sheet(sheet: Worksheet, month: str, year: int, teachers: list, groups: dict):
    """
    Fills out the general statement sheet.

    Args:
        sheet (Worksheet): The sheet to be filled out.
        month (str): The name of the month.
        year (int): The year.
        teachers (list): The list of teachers.
        groups (dict): Dictionary of groups.
    """
    teacher_len = len(teachers)
    group_len = len(groups)

    delete_row_start = 9 + teacher_len + 1
    delete_row_count = 110 - delete_row_start
    end_row = delete_row_start - 1
    
    delete_col_start = 4 + group_len + 1
    delete_col_count = 20 - delete_col_start
    end_col = delete_col_start - 1
    end_col_letter = get_column_letter(end_col)
    end_col_plus_1_letter = get_column_letter(end_col + 1)

    dims = sheet.column_dimensions

    # =============================================================================================================
    # PREPARING THE SHEET BEFORE FILLING

    sheet.unmerge_cells(start_row=113, start_column=3, end_row=113, end_column=20)
    sheet.delete_rows(delete_row_start, amount=delete_row_count)
    for row in range(delete_row_start + 6, delete_row_start + 6 + delete_row_count):
        sheet.row_dimensions[row].hidden = True
    sheet.merge_cells(start_row=delete_row_start + 3, start_column=3, end_row=delete_row_start + 3, end_column=20)

    # Deleting empty columns for nonexistent groups
    if delete_col_count > 0:
        for row in [3, 4, 6, 8, end_row + 4]:
            sheet.unmerge_cells(start_row=row, start_column=3, end_row=row, end_column=20)
        sheet.delete_cols(idx=delete_col_start, amount=delete_col_count)
        for col in range(delete_col_start + 3, delete_col_start + 3 + delete_col_count):
            dims[get_column_letter(col)].hidden = True
        for row in [3, 4, 6, 8, end_row + 4]:
            sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=end_col + 1)
        dims[get_column_letter(delete_col_start + 1)].width = 1.5
        dims[get_column_letter(delete_col_start + 2)].width = 1.5
    
    sheet.print_area = f'C3:{end_col_plus_1_letter}{end_row + 4}'

    # Adding summary formulas by columns and rows
    for col in range(5, end_col + 1):
        col_letter = get_column_letter(col)
        sheet.cell(row=end_row + 1, column=col).value = f'=IF(SUM({col_letter}10:{col_letter}{end_row})=0,"Очікую",SUM({col_letter}10:{col_letter}{end_row}))'
    for row in range(10, end_row + 1):
        sheet.cell(row=row, column=end_col + 1).value = f'=IF(SUM(E{row}:{end_col_letter}{row})=0,"Очікую",SUM(E{row}:{end_col_letter}{row}))'
    sheet.cell(row=end_row + 1, column=end_col + 1).value = f'=IF(SUM({end_col_plus_1_letter}10:{end_col_plus_1_letter}{end_row})=0,"Очікую",SUM({end_col_plus_1_letter}10:{end_col_plus_1_letter}{end_row}))'

    # =============================================================================================================
    # FILLING THE SHEET

    sheet.cell(row=8, column=3).value = f'за {month} {year} р.'

    # Filling the header with group names
    for index, (group_name, group) in enumerate(groups.items()):
        col = 5 + index
        sheet.cell(row=9, column=col).value = group_name

    # Filling teacher data and collecting cell references from other sheets
    for index, teacher in enumerate(teachers):
        row_for_fill = 10 + index
        sheet.cell(row=row_for_fill, column=4).value = teacher['teacher_name']
        for g_element in teacher['groups']:
            group_name = g_element['group_name']
            col_for_fill = groups[group_name] + 4
            column_letter = get_column_letter(g_element['column'])
            row = g_element['row']
            if sheet.cell(row=row_for_fill, column=col_for_fill).value is None:
                sheet.cell(row=row_for_fill, column=col_for_fill).value = f"='{group_name}'!${column_letter}${row}"
            else:
                sheet.cell(row=row_for_fill, column=col_for_fill).value += f" + '{group_name}'!${column_letter}${row}"

def hours_BasedOnTheFirstMonth(info, app_path):
    """
    Creates hours based on the first month.

    Args:
        info (dict): Information about files and months.
        app_path (str): Path to the application.
    Returns:
        dict: Result of the operation with the list of created files.
    """
    answer = {'success': True, 'files': [], 'not_found_subjects': []}
    
    # Creating a directory if it does not exist
    directory_to_save = os.path.dirname(info['file_path'])
    os.makedirs(directory_to_save, exist_ok=True)
    path = f'{app_path}/public/examples/work'

    # =============================================================================================================
    # CREATING HOURS BY MONTHS

    # Iterating through each month and generating the corresponding Excel documents
    for month in info['months']:
        month_name = month['month_name']
        workbook = load_workbook(f'{path}/hours.xlsx')
        step = 1
        for group in month['groups']:
            group_code = group['group_code']
            sheet = workbook[f'Л{step}']
            sheet.title = group_code
            days = {
               'start_day': month['start_day'],
               'end_day': month['end_day'],
               'start_weekday': month['start_weekday']
            }
            filling_out_the_hours_sheet(sheet, month_name, info['year'], group, days)
            step += 1

        path_to_save = f'{directory_to_save}/{month_name.capitalize()}.xlsx'
        data = cleaning_and_save_workbook(workbook, path_to_save)
        if data:
            answer['files'].append(data)

    # =============================================================================================================
    # CREATING GENERAL HOURS FOR THE SEMESTER

    # Generating the general timekeeping file
    workbook = load_workbook(f'{path}/timekeeping.xlsx')
    data = filling_out_the_timekeeping_sheet(workbook, info['months'], info['hours_per_subject'])
    workbook = data['workbook']
    answer['not_found_subjects'] = data['not_found_subjects']

    path_to_save = f'{directory_to_save}/Загальна.xlsx'
    data = cleaning_and_save_workbook(workbook, path_to_save)
    if data:
        answer['files'].append(data)

    return answer

def hours_SummaryOfTeachers(info, appPath):
    """
    Creates a summary statement of teachers.

    Args:
        info (dict): Information for creating the statement.
        appPath (str): Path to the application.
    Returns:
        dict: Result of the operation with the list of created files.
    """
    answer = {'success': True, 'files': []}
    directory_to_save = os.path.dirname(info['file_path'])
    os.makedirs(directory_to_save, exist_ok=True)
    file_name = os.path.basename(info['file_path'])
    path = f'{appPath}/public/examples/work'
    workbook = load_workbook(f'{path}/hours.xlsx')
    month_name = MONTH_NAMES[info['month']]

    # Creating and filling sheets for each group
    step = 1
    groups = {}
    for group in info['groups']:
        group_name = group['group_name']
        sheet = workbook[f'Л{step}']
        sheet.title = group_name
        groups[group_name] = step
        filling_out_the_hours_sheet(sheet, month_name, info['year'], group, info['days'])
        step += 1

    # Formatting the summary sheet "Summary"
    sheet = workbook['Л_Зведена']
    sheet.title = 'Зведена'
    filling_out_the_general_statement_sheet(sheet, month_name, info['year'], info['teachers'], groups)

    path_to_save = f'{directory_to_save}/{file_name}'
    data = cleaning_and_save_workbook(workbook, path_to_save)
    if data:
        answer['files'].append(data)

    return answer