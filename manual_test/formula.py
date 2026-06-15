import openpyxl
from openpyxl.utils import get_column_letter

# --- Использование скрипта ---

# ⚠️ ПЕРЕД ЗАПУСКОМ: 
# 1. Убедитесь, что файл 'my_data.xlsx' существует в той же директории, 
#    или укажите полный путь к нему.
# 2. Убедитесь, что в указанной ячейке на указанном листе есть формула 
#    (например, '=SUM(A1:A10)').
file_name = 'statements14.xlsx'
sheet_to_use = 'Зведена D5'
cell_to_check = ['F36']

def extract_formula(file_path, sheet_name, cell_addresses):
    """
    Открывает файл Excel, извлекает и выводит формулу из указанной ячейки.

    :param file_path: Путь к файлу Excel (например, 'data.xlsx')
    :param sheet_name: Имя листа (например, 'Sheet1')
    :param cell_addresses: Адрес ячейки (например, 'A1' или 'C5')
    """
    try:
        # 1. Загрузка рабочей книги
        print(f"Загрузка файла: {file_path}...")
        workbook = openpyxl.load_workbook(file_path)

        # 2. Выбор рабочего листа
        if sheet_name not in workbook.sheetnames:
            print(f"Ошибка: Лист '{sheet_name}' не найден в файле.")
            print(f"Доступные листы: {workbook.sheetnames}")
            return

        sheet = workbook[sheet_name]
        print(f"Выбран лист: **{sheet_name}**")
        
        for cell_address in cell_addresses:
            # 3. Получение объекта ячейки
            cell = sheet[cell_address]

            # 4. Извлечение и вывод формулы
            formula = cell.value

            if isinstance(formula, str) and formula.startswith('='):
                print(f"\n[{cell_address}]: {formula}")
            elif formula is None:
                print(f"\n[{cell_address}]: ПУСТО")
            else:
                print(f'\n[{cell_address}]: ЗНАЧЕНИЕ "{formula}"')

    except FileNotFoundError:
        print(f"Ошибка: Файл '{file_path}' не найден.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

# Запуск основной функции
extract_formula(file_name, sheet_to_use, cell_to_check)